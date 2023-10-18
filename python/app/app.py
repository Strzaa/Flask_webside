from flask import Flask, request, render_template, redirect, url_for
import openpyxl
from openpyxl import Workbook
from datetime import datetime

def add_link(link, ip, user):
    try:
        workbook = openpyxl.load_workbook("./exel/links.xlsx")
    except FileNotFoundError:
        workbook = Workbook()
        workbook.active.append(['User', 'Link', 'Date', 'IP'])
    sheet = workbook.active

    current_datetime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    sheet.append([user, link, current_datetime, ip])

    workbook.save("./exel/links.xlsx")

    workbook = openpyxl.load_workbook('./exel/goscie.xlsx')
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == user:
            #dont know here
            sheet.cell()

    workbook.save("./exel/goscie.xlsx")

def check_login(username, password=None):
    workbook = openpyxl.load_workbook('./exel/goscie.xlsx')
    sheet = workbook.active

    if password is None:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == username:
                return str(row[2])

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == username and row[1] == password:
            return str(row[2])

    return False

app = Flask(__name__, template_folder='templates')

@app.route('/')
def init():
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']        
        
        sended = check_login(username, password)

        if sended != False:
            if sended == "No":
                return redirect(url_for('sending', user=username))
            else:
                return render_template("done.html")
        else:
            error = 'Invalid Credentials. Please try again.'
    return render_template('login.html', error=error)

@app.route("/send/<user>", methods=["GET", "POST"])
def sending(user):
    if request.method == "GET":
        if check_login(user) == "Yes":
               return render_template("done.html")

    if request.method == "POST":
        link = request.form.get("link")
        ip_adress = request.remote_addr

        add_link(link, ip_adress, user)

    return render_template("send.html")