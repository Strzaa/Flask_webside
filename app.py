from flask import Flask, request, render_template, url_for
import openpyxl
from openpyxl import Workbook
from datetime import datetime

def add_link(user, link, ip):
    workbook = openpyxl.load_workbook("links.xlsx")
    sheet = workbook.active

    current_datetime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    sheet.append([user, link, current_datetime, ip])

    workbook.save("links.xlsx")

def check_user(user):
    try:
        workbook = openpyxl.load_workbook("links.xlsx")
    except FileNotFoundError:
        workbook = Workbook()
        workbook.active.append(['User', 'Link', 'Data', 'IP'])
    sheet = workbook.active

    workbook.save("links.xlsx")

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == user:
            return True
    return False

app = Flask(__name__, template_folder='templates')

@app.route("/send/<user>", methods=["GET", "POST"])
def sending(user):
    if request.method == "POST":
        
        link = request.form.get("link")
        ip_adress = request.remote_addr

        add_link(user, link, ip_adress)

    if check_user(user):
        return render_template("done.html")
    else:
        return render_template("send.html")
